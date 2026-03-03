import streamlit as st
import pandas as pd
import io
import re
import os
import zipfile
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Image, Spacer, Paragraph, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from decimal import Decimal, ROUND_HALF_UP

# ==========================================
# [설정] 리소스 경로 및 상수
# ==========================================
st.set_page_config(page_title="PTI 프로그램", layout="wide")

# 리소스 파일명 (코드와 같은 폴더에 위치해야 함)
LOGO_PATH = "피치트리로고.JPG" 
FONT_PATH = "NanumGothic.ttf"
FONT_BOLD_PATH = "NanumGothic-Bold.ttf"

# Summary 엑셀 컬럼 너비 설정
SUMMARY_WIDTHS = {1: 15, 2: 25, 3: 18, 4: 5, 5: 5, 6: 8, 7: 10, 8: 12, 9: 25, 10: 12, 11: 10, 12: 12, 13: 12, 14: 15, 15: 12, 16: 20}

# Session State 초기화 (KeyError 방지)
for key in ['master_df', 'trace_df', 'review_df']:
    if key not in st.session_state:
        st.session_state[key] = None

# ===================================================================================
# [기능 함수] 스타일 초기화, 파일명 정리, ID 증가, 통화 포맷, 날짜 포맷, 금액 파싱, 스타일 복사
# ===================================================================================
def initialize_styles():
    if os.path.exists(FONT_PATH):
        pdfmetrics.registerFont(TTFont('NanumGothic', FONT_PATH))
    if os.path.exists(FONT_BOLD_PATH):
        pdfmetrics.registerFont(TTFont('NanumGothicBold', FONT_BOLD_PATH))
    else:
        pdfmetrics.registerFont(TTFont('NanumGothicBold', FONT_PATH))

    styles = getSampleStyleSheet()
    custom_styles = {
        'normal': ParagraphStyle(name='Normal', fontName='NanumGothic', fontSize=10, leading=14),
        'title': ParagraphStyle(name='Title', fontName='NanumGothicBold', fontSize=24, alignment=1, leading=26),
        'right': ParagraphStyle(name='Right', fontName='NanumGothic', fontSize=10, alignment=2),
        'center': ParagraphStyle(name='Center', fontName='NanumGothic', fontSize=10, alignment=1),
        'table_bold': ParagraphStyle(name='TableBold', fontName='NanumGothicBold', fontSize=10, alignment=1)
    }
    return custom_styles

def clean_filename(filename):
    val = str(filename).strip()
    if not val or val.lower() in ['nan', 'none']: return ""
    invalid_chars = r'\/:*?"<>|,.' 
    return ''.join(char for char in val if char not in invalid_chars).replace(' ', '_')[:100]

def increment_peachtree_id(pid):
    if pd.isna(pid) or not str(pid).strip(): return ""
    pid_str = str(pid).strip()
    match = re.search(r'(\d+)$', pid_str)
    if match:
        num_str = match.group(1)
        return pid_str[:match.start()] + str(int(num_str) + 1).zfill(len(num_str))
    return pid_str

def format_currency(value):
    try:
        val = float(value)
        return f"₩ {val:,.0f}"
    except: return "₩ 0"

def safe_date_format(value):
    if pd.isna(value) or str(value).strip().lower() in ['', 'nan', 'none']: return ""
    try:
        dt = pd.to_datetime(value, errors='coerce')
        if pd.notnull(dt): return dt.strftime('%Y-%m-%d')
    except: pass
    return str(value)[:10]

def parse_money(value):
    if pd.isna(value): return 0
    num = re.sub(r'[^0-9.]', '', str(value))
    return float(num) if num else 0

def copy_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)

# ==========================================
# [핵심 엔진 1] 정산 로직 (calculate_p)
# ==========================================
def calculate_p(client_db, ip_type, cty, rate, fee_df, db_currency, target_date, manual_usd_rate):
    client_db = str(client_db).strip()
    match_client = fee_df[fee_df['Client'].str.contains(client_db, na=False)]
    if match_client.empty: return 0, "N/A", False

    match = match_client[match_client['IP Type'].str.strip() == ip_type.strip()]
    if match.empty: match = match_client.iloc[[0]]

    target_col = cty if cty in ['CN', 'JP', 'KR'] else 'Our Fees'
    raw_value = str(match.iloc[0][target_col])
    amount = parse_money(raw_value)
    is_usd_fee = 'USD' in raw_value.upper()
    
    actual_rate = rate
    log_label = ""
    if is_usd_fee and db_currency != 'USD':
        actual_rate = manual_usd_rate
        log_label = "(사용자입력 USD환율)"
    else:
        log_label = "(송금환율)"

    val = round(amount * actual_rate) if is_usd_fee else int(amount)
    p_logic = f"{raw_value} x {actual_rate} {log_label} = {val}" if is_usd_fee else f"{val} (원화고정)"
    
    return val, p_logic, is_usd_fee

# ==========================================
# [핵심 엔진 2] PDF 생성 (원본 디자인 100% 동일)
# ==========================================
def generate_pdf_bytes(data):
    styles = initialize_styles()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    client_to = str(data.get('To be invoiced to', '')).strip()
    client_ref = str(data.get('Ref.', '')).strip()
    owner_raw = str(data.get('Owner', '')).strip()
    is_total_invoice = "외" in client_ref and "건" in client_ref
    is_excluded_client = any(name in client_to.upper() for name in ['지평', '세종'])

    # 1. 헤더
    elements.append(Paragraph("서울특별시 서초구 서초대로 356, 905호 (서초동, 서초지웰타워)<br/>Tel.02-6259-2016  Fax.02-6259-2019  E-mail.group@peachtree.co.kr", styles['normal']))
    elements.append(Spacer(1, 15))
    if os.path.exists(LOGO_PATH):
        elements.append(Image(LOGO_PATH, width=1.5*inch, height=1*inch))
    else:
        elements.append(Spacer(1, 50))
    elements.append(Spacer(1, 10))

    doc_main_title = "총    액    청    구    서" if is_total_invoice else "청    구    서"
    elements.append(Paragraph(doc_main_title, styles['title']))
    elements.append(Spacer(1, 20))

    # 2. 정보
    elements.append(Paragraph(f"수신: {client_to}", styles['normal']))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(f"청구번호: {data.get('청구번호', '')}", styles['right']))
    elements.append(Paragraph(f"청구일자: {datetime.now().strftime('%Y-%m-%d')}", styles['right']))
    elements.append(Paragraph(f"송금일자: {safe_date_format(data.get('송금일자'))}", styles['right']))
    elements.append(Spacer(1, 20))

    # 3. 테이블
    table_data = [
        ['대납수수료', '', format_currency(data.get('대납수수료(p)', 0))],
        ['부가세', '', format_currency(data.get('부가세(v)', 0))],
        ['해외비용', '원화환산액', format_currency(data.get('원화환산(w)', 0))],
        ['', '송금수수료', format_currency(data.get('송금수수료(m)', 0))],
        [Paragraph("청구합계", styles['table_bold']), '', Paragraph(format_currency(data.get('청구합계', 0)), styles['table_bold'])]
    ]
    t = Table(table_data, colWidths=[100, 100, 150])
    t.setStyle(TableStyle([
        ('SPAN', (0,0), (1,0)), ('SPAN', (0,1), (1,1)), ('SPAN', (0,2), (0,3)), ('SPAN', (0,4), (1,4)),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (1,1), colors.lightgrey), ('BACKGROUND', (0,2), (1,3), colors.lightgrey), ('BACKGROUND', (0,4), (1,4), colors.lightgrey),
        ('FONTNAME', (0,0), (-1,-1), 'NanumGothic'), ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    # 4. 비고
    def get_clean_val(key):
        v = str(data.get(key, '')).strip() 
        if v.lower() in ['nan', 'none', '']: return None
        return v

    remarks_list = [("[비 고]", True)]
    
    # 1) Your Ref. (Ref.)
    if get_clean_val('Ref.'): 
        remarks_list.append((f"Your Ref.: {data.get('Ref.')}", True))
    
    # 2) Owner (총액청구서가 아닐 때만 출력)
    if not is_total_invoice and owner_raw: 
        remarks_list.append((f"Owner: {owner_raw}", True))
    
    # 3) Total Cost (외화합계) - 통화 기호와 함께 출력
    curr = get_clean_val('통화')
    total_foreign = get_clean_val('외화합계(t)')
    if total_foreign:
        # 외화가 숫자인 경우 천단위 콤마 포맷팅 (선택 사항, 필요 없으면 그냥 total_foreign 사용)
        try:
            formatted_foreign = f"{float(total_foreign):,.2f}"
        except:
            formatted_foreign = total_foreign
        remarks_list.append((f"Total Cost: {curr or ''} {formatted_foreign}", True))
        
    # 4) Country Code (Cty)
    if get_clean_val('Cty'): 
        remarks_list.append((f"Country Code: {data.get('Cty')}", True))
    
    # 한 칸 띄우기 및 할증 문구
    remarks_list.append(("", True))
    if not is_excluded_client:
        remarks_list.append(("청구일자로부터 6영업일부터 청구 전액에 대해 일할 할증됩니다.(월이자1% 기준)", True))
    
    # PDF에 반영
    for text, condition in remarks_list:
        if condition:
            elements.append(Paragraph(text, styles['normal']))

    elements.append(Spacer(1, 15))
    elements.append(Paragraph('입금계좌: 하나은행(189-910056-22204, 예금주: 주식회사 피치트리)', styles['center']))
    
    doc.build(elements)
    return buffer.getvalue()
    
# ==========================================
# [핵심 엔진 3] Summary 생성
# ==========================================
def generate_summary_bytes(rows_list, template_obj):
    template_obj.seek(0)
    wb = openpyxl.load_workbook(template_obj)
    ws = wb.active
    for i, row_vals in enumerate(rows_list):
        for j, val in enumerate(row_vals):
            cell = ws.cell(row=2 + i, column=j + 1)
            cell.value = val
            copy_style(ws.cell(row=2, column=j + 1), cell)
    for col_idx, width in SUMMARY_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ==========================================
# [메인 화면] UI 및 정산 실행
# ==========================================
st.title("💵 PTI 프로그램 (PeachTree Invoicing)")

st.sidebar.header("📁 Step 1: 파일 업로드")
up_db = st.sidebar.file_uploader("데이터베이스 CSV", type="csv")
up_fee = st.sidebar.file_uploader("수수료표 CSV ", type="csv")
up_temp = st.sidebar.file_uploader("Summary 템플릿 XLSX ", type="xlsx")

if up_db and up_fee:
    df_input = pd.read_csv(up_db)
    df_fee = pd.read_csv(up_fee)
    
    # 이종 통화 감지
    needs_usd_rate = False
    mismatch = df_input[df_input['currency'].str.upper() != 'USD']
    if not mismatch.empty:
        needs_usd_rate = True
        sample_date = mismatch['송금일자'].iloc[0]

    manual_usd_rate = 0.0
    if needs_usd_rate:
        st.warning(f"⚠️ 이종 통화 감지 - 송수신 통화와 수수료 통화가 다릅니다! 송금일: {sample_date}의 USD(송금받을 때) 환율을 입력해주세요.")
        manual_usd_rate = st.number_input("USD 환율(송금받을때) 입력:", value=0.0, step=0.1)

    if st.button("수수료 정산"):
        if needs_usd_rate and manual_usd_rate == 0:
            st.error("환율을 입력하세요!")
        else:
            final_rows, calc_traces, review_rows = [], [], []
            for gcd, group in df_input.groupby("Invoice ID", sort=False):
                client_db = str(group["Client"].iloc[0]).strip()
                r, f_date, db_curr = group["송금환율"].iloc[0], group["송금일자"].iloc[0], str(group["currency"].iloc[0]).strip().upper()
                total_k, n = group["송금수수료"].fillna(0).max(), len(group)
                m = int(total_k / n) if n > 0 else 0
                
                last_pt_id = ""
                for _, row in group.iterrows():
                    ip_map = {"p": "Patent", "u": "Utility Model", "d": "Design", "t": "TradeMark"}
                    ip_type = ip_map.get(str(row["category"]).lower().strip(), "Patent")
                    p, p_logic, _ = calculate_p(client_db, ip_type, row["Cty"], r, df_fee, db_curr, f_date, manual_usd_rate)
                    v, w = round(p * 0.1), int(row["total"] * r)
                    bill_row, last_pt_id = p + v + w + m, row["PeachTree ID"]

                    p_fee_match = re.search(r'(?:USD\s?)?([\d\.,]+)', p_logic)
                    p_fee_only = p_fee_match.group(1).replace(',', '') if p_fee_match else ""
                    # 1. 대납수수료 기준 추출
                    p_basis_match = re.search(r'\((.*?)\)', p_logic)
                    p_basis = p_basis_match.group(1) if p_basis_match else ""
                    
                    # [수정] 사용자입력 USD환율 케이스까지 완벽 처리
                    is_usd_logic = p_basis in ["송금환율", "사용자입력 USD환율"]
                    p_currency = row["currency"] if is_usd_logic else ("KRW" if p_basis == "원화고정" else "")
                    # 사용자 입력일 땐 manual_usd_rate를, 아닐 땐 r(송금환율)을 적용
                    p_rate = (manual_usd_rate if "사용자입력" in p_basis else r) if is_usd_logic else (1 if p_basis == "원화고정" else "")

                    final_rows.append({"code": gcd, "송금확인번호": row.get("송금확인번호", ""), "Ref.": row["Client ID"], "청구번호": last_pt_id, "Cty": row["Cty"], "yr": row["yr"], "통화": row["currency"], "환율": r, "외화합계(t)": row["total"], "To be invoiced to": client_db, "대납수수료(p)": p, "부가세(v)": v, "원화환산(w)": w, "송금수수료(m)": m, "청구합계": bill_row, "송금일자": f_date, "Owner": row["owner"]})
                    calc_traces.append({"Invoice ID": gcd, "Ref.": row["Client ID"], "청구번호": last_pt_id, "1. 대납수수료(p)": p_logic, "2. 부가세(v)": f"{p} x 0.1 = {v}", "3. 원화환산(w)": f"{row['total']} x {r} = {w}", "4. 송금수수료(m)": f"{total_k}/{n}={m}", "최종 합계": f"{p}+{v}+{w}+{m}={bill_row}"})
                    
                    # 📝 검토용 데이터 (일반 행)
                    review_rows.append({
                        "Ref.": row["Client ID"], 
                        "청구번호": last_pt_id, 
                        "p_기준": p_basis, "p_통화": p_currency, "p_수수료": p_fee_only, "p_적용환율": p_rate, "p_계산": p, "p_일치": "TRUE",
                        "v_비율": "10%", "v_계산": v, "v_일치": "TRUE",
                        "w_통화": row["currency"], "w_적용환율": r, "w_외화": row["total"], "w_원화": w, "w_일치": "TRUE",
                        "m_금액": total_k, "m_수량": n, "m_결과": m, "m_일치": "TRUE"
                    })

                if n > 1:
                    last = final_rows[-1]
                    p_s, v_s, group_t_sum = sum(r['대납수수료(p)'] for r in final_rows[-n:]), sum(r['부가세(v)'] for r in final_rows[-n:]), group["total"].sum()
                    w_s, new_pt_id = int(group_t_sum * r), increment_peachtree_id(last_pt_id)
                    bill_grand = p_s + v_s + w_s + total_k
                    
                    final_rows.append({"code": f"[{gcd} 합계]", "송금확인번호": group["송금확인번호"].iloc[0] if "송금확인번호" in group.columns else "", "Ref.": f"{group['Client ID'].iloc[0]} 외 {n-1}건", "청구번호": new_pt_id, "Cty": "", "yr": "", "통화": last["통화"], "환율": r, "외화합계(t)": group_t_sum, "To be invoiced to": client_db, "대납수수료(p)": p_s, "부가세(v)": v_s, "원화환산(w)": w_s, "송금수수료(m)": total_k, "청구합계": bill_grand, "송금일자": f_date, "Owner": ""})
                    calc_traces.append({"Invoice ID": f"[{gcd} 합계]", "Ref.": f"{group['Client ID'].iloc[0]} 외 {n-1}건", "청구번호": new_pt_id, "1. 대납수수료(p)": f"p합산={p_s}", "2. 부가세(v)": f"v합산={v_s}", "3. 원화환산(w)": f"{group_t_sum}x{r}={w_s}", "4. 송금수수료(m)": f"총수수료={total_k}", "최종 합계": f"{p_s}+{v_s}+{w_s}+{total_k}={bill_grand}"})
                    
                    # 📝 검토용 데이터 (합계 행)
                    review_rows.append({
                        "Ref.": f"{group['Client ID'].iloc[0]} 외 {n-1}건",
                        "청구번호": new_pt_id, # [수정] 합계 행 전용 ID로 변경
                        "p_기준": "", "p_통화": "", "p_수수료": "", "p_적용환율": "", "p_계산": p_s, "p_일치": "TRUE",
                        "v_비율": "10%", "v_계산": v_s, "v_일치": "TRUE",
                        "w_통화": last["통화"], "w_적용환율": r, "w_외화": group_t_sum, "w_원화": w_s, "w_일치": "TRUE",
                        "m_금액": "", "m_수량": "", "m_결과": total_k, "m_일치": "TRUE"
                    })
            master_df = pd.DataFrame(final_rows)
            trace_df = pd.DataFrame(calc_traces)
            
            review_df = pd.DataFrame(review_rows)
            
            multi_cols = [
                (" ", "Ref."),
                (" ", "청구번호"),
                ("1. 대납수수료(p)", "대납수수료 기준"), ("1. 대납수수료(p)", "통화"), ("1. 대납수수료(p)", "수수료"), 
                ("1. 대납수수료(p)", "적용환율"), ("1. 대납수수료(p)", "대납수수료 계산"), ("1. 대납수수료(p)", "일치여부"),
                ("2. 부가세(v)", "부가세율"), ("2. 부가세(v)", "부가세 계산"), ("2. 부가세(v)", "일치여부"),
                ("3. 원화환산(w)", "통화"), ("3. 원화환산(w)", "적용환율"), ("3. 원화환산(w)", "외화합계"), 
                ("3. 원화환산(w)", "원화환산"), ("3. 원화환산(w)", "일치여부"),
                ("4. 송금수수료(m)", "금액"), ("4. 송금수수료(m)", "청구수량"), ("4. 송금수수료(m)", "결과"), 
                ("4. 송금수수료(m)", "일치여부")
            ]
            review_df.columns = pd.MultiIndex.from_tuples(multi_cols)
            
            st.session_state['master_df'] = master_df
            st.session_state['trace_df'] = trace_df
            st.session_state['review_df'] = review_df
            st.success("✅ 정산 완료!")

# 결과 출력 및 다운로드
if st.session_state['master_df'] is not None:
    df_res = st.session_state['master_df']
    df_review = st.session_state['review_df']
    st.dataframe(df_res)
    c1, c2, c3 = st.columns(3)
    with c1:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df_res.to_excel(writer, index=False, sheet_name='Summary')
            st.session_state['trace_df'].to_excel(writer, index=False, sheet_name='Calculation_Trace')
            df_review_data = df_review.copy()
            df_review_data.columns = [c[1] for c in df_review.columns] # 2단 제목만 추출
            df_review_data.to_excel(writer, index=False, sheet_name='검토', startrow=1) # 2행부터 데이터 시작
            
            workbook = writer.book
            ws_review = writer.sheets['검토']
            
            header_fmt = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
            
            ws_review.merge_range('A1:B1', ' ', header_fmt) 
            ws_review.merge_range('C1:H1', '1. 대납수수료(p)', header_fmt)
            ws_review.merge_range('I1:K1', '2. 부가세(v)', header_fmt)
            ws_review.merge_range('L1:P1', '3. 원화환산(w)', header_fmt)
            ws_review.merge_range('Q1:T1', '4. 송금수수료(m)', header_fmt)
            
            for col_num, value in enumerate(df_review_data.columns.values):
                ws_review.write(1, col_num, value, header_fmt)
            gray_fmt = workbook.add_format({
                'bg_color': '#F2F2F2', 
                'border': 1, 
                'align': 'center', 
                'valign': 'vcenter'
            })
            # 2. 사이드 텍스트용 (왼쪽 정렬 유지)
            side_text_fmt = workbook.add_format({
                'bg_color': '#F2F2F2', 
                'border': 1, 
                'align': 'left', 
                'valign': 'vcenter'
            })
            # 3. 헤더 굵게
            header_fmt = workbook.add_format({
                'bold': True, 
                'bg_color': '#F2F2F2', 
                'border': 1, 
                'align': 'center', 
                'valign': 'vcenter'
            })
            # 4단계: 동적 수식 및 실시간 검증 적용 (3행부터) [cite: 2026-03-03]
            # p_계산(G열) = 수수료 * 환율
            # p_일치(H열) = G열 값과 Summary 시트의 대납수수료(K열) 비교
            rate_formula = '=IFERROR(IF(C{row}="원화고정", 1, IF(C{row}="송금환율", _xlfn.XLOOKUP(B{row}, Summary!$D:$D, Summary!$H:$H, ""), {p_rate})), "")'
            calc_formula = '=ROUND(E{row} * F{row}, 0)'
            
            # 실시간 검증 수식 (Summary 시트의 각 항목과 비교) [cite: 2026-03-03]
            check_p = '=G{row} = _xlfn.XLOOKUP(B{row}, Summary!$D:$D, Summary!$K:$K)' # 대납수수료 일치여부 [cite: 2026-03-03]
            check_v = '=J{row} = _xlfn.XLOOKUP(B{row}, Summary!$D:$D, Summary!$L:$L)' # 부가세 일치여부 [cite: 2026-03-03]
            check_w = '=O{row} = _xlfn.XLOOKUP(B{row}, Summary!$D:$D, Summary!$M:$M)' # 원화환산 일치여부 [cite: 2026-03-03]
            check_m = '=S{row} = _xlfn.XLOOKUP(B{row}, Summary!$D:$D, Summary!$N:$N)' # 송금수수료 일치여부 [cite: 2026-03-03]
            
            for i in range(len(df_review_data)):
                row_idx = i + 2 
                row_num = row_idx + 1 
                
                # A, B열 배경색 유지
                ws_review.write(row_idx, 0, df_review_data.iloc[i, 0], side_gray_fmt)
                ws_review.write(row_idx, 1, df_review_data.iloc[i, 1], side_gray_fmt)
                
                ref_val = str(df_review_data.iloc[i, 0])
                if "합계" not in ref_val:
                    # 1. 적용환율 및 계산 수식
                    p_rate_val = df_review_data.iloc[i, 5]
                    ws_review.write_formula(row_idx, 5, rate_formula.format(row=row_num, p_rate=p_rate_val or 1), data_fmt)
                    ws_review.write_formula(row_idx, 6, calc_formula.format(row=row_num), won_fmt)
                    
                    # 2. 실시간 일치여부 검증 수식 (글자 대신 수식 입력) [cite: 2026-03-03]
                    ws_review.write_formula(row_idx, 7, check_p.format(row=row_num), data_fmt)  # H열
                    ws_review.write_formula(row_idx, 10, check_v.format(row=row_num), data_fmt) # K열
                    ws_review.write_formula(row_idx, 15, check_w.format(row=row_num), data_fmt) # P열
                    ws_review.write_formula(row_idx, 19, check_m.format(row=row_num), data_fmt) # T열
                else:
                    # 합계 행은 계산된 값 그대로 기록
                    ws_review.write(row_idx, 6, df_review_data.iloc[i, 6], won_fmt)
                    ws_review.write(row_idx, 7, "TRUE", data_fmt) # 합계는 수동 확인

                # 부가세(J), 원화환산(O), 결과(S) 데이터 기록
                for col_idx in [9, 14, 18]: 
                    ws_review.write(row_idx, col_idx, df_review_data.iloc[i, col_idx], won_fmt)

            won_fmt = workbook.add_format({'num_format': '₩#,##0', 'align': 'right'})
            writer.sheets['Summary'].set_column('J:N', 15, won_fmt)
            writer.sheets['Calculation_Trace'].set_column('A:H', 28)
            ws_review.set_column('A:T', 15) 
        st.download_button("📂 1단계: 메인 엑셀 다운로드.2", output.getvalue(), "PTI_main_result.xlsx")
    with c2:
        if st.button("📄 2단계: 비용청구서 PDF 일괄 생성"):
            pdf_zip = io.BytesIO()
            with zipfile.ZipFile(pdf_zip, "w") as zf:
                for _, r in df_res.iterrows():
                    is_t = "외" in str(r['Ref.'])
                    pre = "[총액청구서]" if is_t else (f"[{clean_filename(r['Owner'])}]" if r['Owner'] else "")
                    zf.writestr(f"{pre}{clean_filename(r['To be invoiced to'])}_{clean_filename(r['Ref.'])}.pdf", generate_pdf_bytes(r))
            st.download_button("PDF 다운로드", pdf_zip.getvalue(), "PTI_Invoices.zip")
    with c3:
        if up_temp and st.button("📊 3단계: Summary 일괄 생성"):
            sum_zip = io.BytesIO()
            with zipfile.ZipFile(sum_zip, "w") as zf:
                df_res['pure'] = df_res['code'].apply(lambda x: str(x).split(']')[0].replace('[', '').split()[0] if '[' in str(x) else str(x))
                for _, gp in df_res.groupby('pure'):
                    g_list = gp.drop(columns=['pure']).values.tolist()
                    zf.writestr(f"Summary_{str(g_list[0][8])}_{str(g_list[-1][1])}.xlsx".replace('/', '_'), generate_summary_bytes(g_list, up_temp))

            st.download_button("Summary 엑셀 다운로드", sum_zip.getvalue(), "Summaries.zip")
